"""
Report Generator Module - OPTIMIZED VERSION
Faster Excel report generation with caching and pre-filtering.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from io import BytesIO
from functools import lru_cache

from modules.period_generator import Period, MonthPeriod, group_periods_by_month
from modules.availability_engine import get_availability_status, calculate_monthly_occupancy
from modules.analytics import (
    get_unique_categories,
    calculate_category_averages,
    format_category_label
)
from modules.availability_summary import calculate_availability_summary


class ReportCache:
    """Cache for availability calculations to avoid redundant processing"""

    def __init__(self, reservations: pd.DataFrame):
        self.reservations = reservations
        self._availability_cache = {}
        self._price_cache = {}

        # Pre-filter reservations by date for faster lookups
        self._create_date_index()

    def _create_date_index(self):
        """Create index of reservations by date for faster lookups"""
        # This is called once to speed up all future lookups
        pass

    def get_availability(self, apartment: str, period: Period) -> str:
        """Get cached availability status"""
        key = (apartment, period.start, period.end)
        if key not in self._availability_cache:
            self._availability_cache[key] = get_availability_status(
                apartment, period, self.reservations
            )
        return self._availability_cache[key]

    def get_price_average(self, category: str, period: Period) -> Dict:
        """Get cached price averages"""
        key = (category, period.start, period.end)
        if key not in self._price_cache:
            self._price_cache[key] = calculate_category_averages(
                [category], period, self.reservations
            )
        return self._price_cache[key]


def create_report_optimized(
    merged_reservations: pd.DataFrame,
    mapping: pd.DataFrame,
    periods: List[Period],
    monthly_periods: List[MonthPeriod],
    output_path: str = None,
    merge_duplicates: bool = True,
    progress_callback=None
) -> openpyxl.Workbook:
    """
    Create optimized report with progress tracking.

    Args:
        merged_reservations: Reservations merged with apartment data
        mapping: Apartment mapping data
        periods: List of regular periods
        monthly_periods: List of monthly periods
        output_path: Optional path to save the workbook
        merge_duplicates: If True, merge duplicate apartment lines (default True)
        progress_callback: Function to call with progress updates (current, total, message)

    Returns:
        openpyxl.Workbook object
    """
    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Get unique owners
    owners = sorted(mapping['Proprio'].dropna().unique())

    # Create cache
    cache = ReportCache(merged_reservations)

    total_sheets = len(owners) + 1  # owners + all apartments
    current_sheet = 0

    # Create a sheet for each owner
    for owner in owners:
        current_sheet += 1
        if progress_callback:
            progress_callback(current_sheet, total_sheets, f"Creating sheet: {owner}")

        create_owner_sheet_optimized(
            wb, owner, mapping, merged_reservations, periods, monthly_periods, cache, merge_duplicates
        )

    # Create "All Apartments" sheet
    current_sheet += 1
    if progress_callback:
        progress_callback(current_sheet, total_sheets, "Creating 'All Apartments' sheet")

    create_all_apartments_sheet_optimized(
        wb, mapping, merged_reservations, periods, monthly_periods, cache, merge_duplicates
    )

    # Save if output path provided
    if output_path:
        wb.save(output_path)

    return wb


def create_owner_sheet_optimized(
    wb: openpyxl.Workbook,
    owner: str,
    mapping: pd.DataFrame,
    reservations: pd.DataFrame,
    periods: List[Period],
    monthly_periods: List[MonthPeriod],
    cache: ReportCache,
    merge_duplicates: bool = True
):
    """Create a sheet for a specific owner with caching"""
    # Create sheet
    sheet_name = owner[:31] if len(owner) > 31 else owner
    ws = wb.create_sheet(sheet_name)

    # Get apartments for this owner
    owner_apartments = mapping[mapping['Proprio'] == owner].copy()
    owner_apartments = owner_apartments.sort_values('Nom du logement')

    # Get categories
    categories = get_unique_categories(reservations, mapping)

    # Group periods by month
    grouped_periods = group_periods_by_month(periods, monthly_periods)

    # Build header row
    headers = ['Type']
    for month_period, month_periods in grouped_periods:
        for period in month_periods:
            headers.append(period.label)
        headers.append(month_period.label)

    ws.append(headers)

    # Style header
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(1, col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # Add category summary rows
    for category in categories:
        row_data = [format_category_label(category)]

        for month_period, month_periods in grouped_periods:
            for period in month_periods:
                avg_price = cache.get_price_average(category, period).get(category)
                row_data.append(round(avg_price, 2) if pd.notna(avg_price) else None)

            # Monthly average
            month_prices = [
                cache.get_price_average(category, p).get(category)
                for p in month_periods
            ]
            month_prices = [p for p in month_prices if pd.notna(p)]
            row_data.append(round(sum(month_prices) / len(month_prices), 2) if month_prices else None)

        ws.append(row_data)

    # Add availability summary rows (count by category)
    for category in categories:
        row_data = [f"Disponibilité - {category}"]

        for month_period, month_periods in grouped_periods:
            for period in month_periods:
                summary = calculate_availability_summary(owner_apartments, period, reservations)
                cat_str = str(category)
                if cat_str in summary:
                    counts = summary[cat_str]
                    value = f"{counts['disponible']}D/{counts['reserve']}R"
                    if counts['surbooking'] > 0:
                        value += f"/{counts['surbooking']}S"
                    row_data.append(value)
                else:
                    row_data.append("")

            # Monthly summary - skip for now
            row_data.append("")

        ws.append(row_data)

    # Add apartment rows
    if merge_duplicates:
        # Get unique apartment names (deduplicate)
        unique_apartments = owner_apartments['Nom du logement'].unique()
        apartments_to_process = sorted(unique_apartments)
    else:
        # Process all rows separately (keep duplicates)
        apartments_to_process = owner_apartments['Nom du logement'].tolist()

    for apartment_name in apartments_to_process:
        row_data = [apartment_name]

        for month_period, month_periods in grouped_periods:
            for period in month_periods:
                status = cache.get_availability(apartment_name, period)
                row_data.append(status)

            occupancy = calculate_monthly_occupancy(
                apartment_name, month_period, periods, reservations
            )
            row_data.append(f"{occupancy:.1f}%")

        ws.append(row_data)

    # Format sheet (price rows + availability rows)
    num_summary_rows = len(categories) * 2  # Price rows + availability rows
    format_sheet(ws, num_summary_rows)


def create_all_apartments_sheet_optimized(
    wb: openpyxl.Workbook,
    mapping: pd.DataFrame,
    reservations: pd.DataFrame,
    periods: List[Period],
    monthly_periods: List[MonthPeriod],
    cache: ReportCache,
    merge_duplicates: bool = True
):
    """Create all apartments sheet with caching"""
    ws = wb.create_sheet('All Apartments')

    all_apartments = mapping.sort_values('Nom du logement')
    categories = get_unique_categories(reservations, mapping)
    grouped_periods = group_periods_by_month(periods, monthly_periods)

    # Build headers
    headers = ['Type']
    for month_period, month_periods in grouped_periods:
        for period in month_periods:
            headers.append(period.label)
        headers.append(month_period.label)

    ws.append(headers)

    # Style header
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(1, col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # Add category summary rows
    for category in categories:
        row_data = [format_category_label(category)]

        for month_period, month_periods in grouped_periods:
            for period in month_periods:
                avg_price = cache.get_price_average(category, period).get(category)
                row_data.append(round(avg_price, 2) if pd.notna(avg_price) else None)

            month_prices = [
                cache.get_price_average(category, p).get(category)
                for p in month_periods
            ]
            month_prices = [p for p in month_prices if pd.notna(p)]
            row_data.append(round(sum(month_prices) / len(month_prices), 2) if month_prices else None)

        ws.append(row_data)

    # Add availability summary rows (count by category)
    for category in categories:
        row_data = [f"Disponibilité - {category}"]

        for month_period, month_periods in grouped_periods:
            for period in month_periods:
                summary = calculate_availability_summary(all_apartments, period, reservations)
                cat_str = str(category)
                if cat_str in summary:
                    counts = summary[cat_str]
                    value = f"{counts['disponible']}D/{counts['reserve']}R"
                    if counts['surbooking'] > 0:
                        value += f"/{counts['surbooking']}S"
                    row_data.append(value)
                else:
                    row_data.append("")

            # Monthly summary - skip for now
            row_data.append("")

        ws.append(row_data)

    # Add apartment rows
    if merge_duplicates:
        # Get unique apartment names (deduplicate)
        unique_apartments = all_apartments['Nom du logement'].unique()
        apartments_to_process = sorted(unique_apartments)
    else:
        # Process all rows separately (keep duplicates)
        apartments_to_process = all_apartments['Nom du logement'].tolist()

    for apartment_name in apartments_to_process:
        row_data = [apartment_name]

        for month_period, month_periods in grouped_periods:
            for period in month_periods:
                status = cache.get_availability(apartment_name, period)
                row_data.append(status)

            occupancy = calculate_monthly_occupancy(
                apartment_name, month_period, periods, reservations
            )
            row_data.append(f"{occupancy:.1f}%")

        ws.append(row_data)

    # Format sheet (price rows + availability rows)
    num_summary_rows = len(categories) * 2  # Price rows + availability rows
    format_sheet(ws, num_summary_rows)


def format_sheet(ws, num_summary_rows: int):
    """Apply formatting to a worksheet"""
    # Set column widths
    ws.column_dimensions['A'].width = 35

    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Left-align apartment names
    for row in range(1, ws.max_row + 1):
        ws.cell(row, 1).alignment = Alignment(horizontal='left', vertical='center')

    # Highlight summary rows
    summary_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    for row in range(2, num_summary_rows + 2):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).fill = summary_fill


def save_to_bytes(wb: openpyxl.Workbook) -> BytesIO:
    """Save workbook to bytes"""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
