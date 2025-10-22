"""
Report Generator Module
Generate Excel reports with availability and pricing data.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from io import BytesIO

from modules.period_generator import Period, MonthPeriod, group_periods_by_month
from modules.availability_engine import get_availability_status, calculate_monthly_occupancy
from modules.analytics import (
    get_unique_categories,
    calculate_category_averages,
    format_category_label
)


def create_report(
    merged_reservations: pd.DataFrame,
    mapping: pd.DataFrame,
    periods: List[Period],
    monthly_periods: List[MonthPeriod],
    output_path: str = None
) -> openpyxl.Workbook:
    """
    Create the complete availability report workbook.

    Args:
        merged_reservations: Reservations merged with apartment data
        mapping: Apartment mapping data
        periods: List of regular periods
        monthly_periods: List of monthly periods
        output_path: Optional path to save the workbook

    Returns:
        openpyxl.Workbook object
    """
    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Get unique owners
    owners = sorted(mapping['Proprio'].dropna().unique())

    print(f"\nGenerating report with {len(owners)} owner sheets...")

    # Create a sheet for each owner
    for owner in owners:
        print(f"  Creating sheet for: {owner}")
        create_owner_sheet(
            wb, owner, mapping, merged_reservations, periods, monthly_periods
        )

    # Create "All Apartments" sheet
    print("  Creating 'All Apartments' sheet")
    create_all_apartments_sheet(
        wb, mapping, merged_reservations, periods, monthly_periods
    )

    # Save if output path provided
    if output_path:
        wb.save(output_path)
        print(f"\n✓ Report saved to: {output_path}")

    return wb


def create_owner_sheet(
    wb: openpyxl.Workbook,
    owner: str,
    mapping: pd.DataFrame,
    reservations: pd.DataFrame,
    periods: List[Period],
    monthly_periods: List[MonthPeriod]
):
    """
    Create a sheet for a specific owner.

    Args:
        wb: Workbook to add sheet to
        owner: Owner name
        mapping: Apartment mapping data
        reservations: Merged reservation data
        periods: List of regular periods
        monthly_periods: List of monthly periods
    """
    # Create sheet (trim owner name to max 31 chars for Excel)
    sheet_name = owner[:31] if len(owner) > 31 else owner
    ws = wb.create_sheet(sheet_name)

    # Get apartments for this owner
    owner_apartments = mapping[mapping['Proprio'] == owner].copy()
    owner_apartments = owner_apartments.sort_values('Nom du logement')

    # Get categories
    categories = get_unique_categories(reservations, mapping)

    # Group periods by month for column organization
    grouped_periods = group_periods_by_month(periods, monthly_periods)

    # Build header row
    headers = ['Type']
    for month_period, month_periods in grouped_periods:
        # Add period columns
        for period in month_periods:
            headers.append(period.label)
        # Add monthly summary column
        headers.append(month_period.label)

    # Write headers
    ws.append(headers)

    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(1, col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # Add category summary rows
    for category in categories:
        row_data = [format_category_label(category)]

        for month_period, month_periods in grouped_periods:
            # Add prices for each period
            for period in month_periods:
                avg_price = calculate_category_averages(
                    [category], period, reservations
                ).get(category)

                if pd.notna(avg_price):
                    row_data.append(round(avg_price, 2))
                else:
                    row_data.append(None)

            # Add monthly average
            month_prices = []
            for period in month_periods:
                avg_price = calculate_category_averages(
                    [category], period, reservations
                ).get(category)
                if pd.notna(avg_price):
                    month_prices.append(avg_price)

            if month_prices:
                row_data.append(round(sum(month_prices) / len(month_prices), 2))
            else:
                row_data.append(None)

        ws.append(row_data)

    # Add individual apartment rows
    for _, apartment_row in owner_apartments.iterrows():
        apartment_name = apartment_row['Nom du logement']
        row_data = [apartment_name]

        for month_period, month_periods in grouped_periods:
            # Add availability status for each period
            for period in month_periods:
                status = get_availability_status(apartment_name, period, reservations)
                row_data.append(status)

            # Add monthly occupancy
            occupancy = calculate_monthly_occupancy(
                apartment_name, month_period, periods, reservations
            )
            row_data.append(f"{occupancy:.1f}%")

        ws.append(row_data)

    # Format the sheet
    format_sheet(ws, len(categories))


def create_all_apartments_sheet(
    wb: openpyxl.Workbook,
    mapping: pd.DataFrame,
    reservations: pd.DataFrame,
    periods: List[Period],
    monthly_periods: List[MonthPeriod]
):
    """
    Create a sheet with all apartments combined.

    Args:
        wb: Workbook to add sheet to
        mapping: Apartment mapping data
        reservations: Merged reservation data
        periods: List of regular periods
        monthly_periods: List of monthly periods
    """
    ws = wb.create_sheet('All Apartments')

    # Get all apartments
    all_apartments = mapping.sort_values('Nom du logement')

    # Get categories
    categories = get_unique_categories(reservations, mapping)

    # Group periods by month
    grouped_periods = group_periods_by_month(periods, monthly_periods)

    # Build header row
    headers = ['Type']
    for month_period, month_periods in grouped_periods:
        for period in month_periods:
            headers.append(period.label)
        headers.append(month_period.label)

    ws.append(headers)

    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(1, col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # Add category summary rows
    for category in categories:
        row_data = [format_category_label(category)]

        for month_period, month_periods in grouped_periods:
            for period in month_periods:
                avg_price = calculate_category_averages(
                    [category], period, reservations
                ).get(category)

                if pd.notna(avg_price):
                    row_data.append(round(avg_price, 2))
                else:
                    row_data.append(None)

            # Monthly average
            month_prices = []
            for period in month_periods:
                avg_price = calculate_category_averages(
                    [category], period, reservations
                ).get(category)
                if pd.notna(avg_price):
                    month_prices.append(avg_price)

            if month_prices:
                row_data.append(round(sum(month_prices) / len(month_prices), 2))
            else:
                row_data.append(None)

        ws.append(row_data)

    # Add individual apartment rows
    for _, apartment_row in all_apartments.iterrows():
        apartment_name = apartment_row['Nom du logement']
        row_data = [apartment_name]

        for month_period, month_periods in grouped_periods:
            for period in month_periods:
                status = get_availability_status(apartment_name, period, reservations)
                row_data.append(status)

            occupancy = calculate_monthly_occupancy(
                apartment_name, month_period, periods, reservations
            )
            row_data.append(f"{occupancy:.1f}%")

        ws.append(row_data)

    # Format the sheet
    format_sheet(ws, len(categories))


def format_sheet(ws, num_summary_rows: int):
    """
    Apply formatting to a worksheet.

    Args:
        ws: Worksheet to format
        num_summary_rows: Number of summary rows (category averages) at the top
    """
    # Set column widths
    ws.column_dimensions['A'].width = 35  # Apartment name column

    # Auto-size other columns
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Left-align the apartment names in column A
    for row in range(1, ws.max_row + 1):
        ws.cell(row, 1).alignment = Alignment(horizontal='left', vertical='center')

    # Highlight summary rows
    summary_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    for row in range(2, num_summary_rows + 2):  # +2 because row 1 is header, rows 2+ are summaries
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).fill = summary_fill


def save_to_bytes(wb: openpyxl.Workbook) -> BytesIO:
    """
    Save workbook to bytes (for Streamlit download).

    Args:
        wb: Workbook to save

    Returns:
        BytesIO object containing the Excel file
    """
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


if __name__ == "__main__":
    # Test the module
    from pathlib import Path
    from modules.data_loader import load_and_prepare_data
    from modules.period_generator import generate_periods, generate_monthly_periods
    from datetime import date

    print("\n" + "="*60)
    print("Testing Report Generator")
    print("="*60)

    # Load sample data
    base_path = Path(__file__).parent.parent
    mapping_file = base_path / "Fichier de mapping par appartement.xlsx"
    reservations_file = base_path / "Liste des réservations-22-10-2025-31-10-2025.csv"

    merged, mapping = load_and_prepare_data(str(mapping_file), str(reservations_file))

    # Generate periods for October only (to keep test small)
    start = date(2025, 10, 22)
    end = date(2025, 10, 31)
    periods = generate_periods(start, end, period_days=3)
    monthly = generate_monthly_periods(start, end)

    print(f"\nGenerating test report for {start} to {end}")
    print(f"  {len(periods)} periods, {len(monthly)} months")

    # Create report
    output_file = base_path / "test_report.xlsx"
    wb = create_report(merged, mapping, periods, monthly, str(output_file))

    print("\n" + "="*60 + "\n")
